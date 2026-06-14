"""Mine DO-wave + driver-habit patterns from all 2026 schedules for placement policy."""
from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

SCHED = Path(r"C:\Users\megap\Desktop\Schedule for 2026")
OUT = Path(__file__).resolve().parent / "placement-policy-mined.json"

TRIP = re.compile(r"^1-\d+-([ABC])$", re.I)
META = re.compile(r"^__FS")
TIME = re.compile(r"(\d{1,2}):(\d{2})\s*(AM|PM)?", re.I)


def parse_time(cell):
    if cell is None:
        return None
    if hasattr(cell, "hour") and hasattr(cell, "minute"):
        return cell.hour * 60 + cell.minute
    m = TIME.search(str(cell))
    if not m:
        return None
    h, mi = int(m.group(1)), int(m.group(2))
    ap = (m.group(3) or "").upper()
    if ap == "PM" and h != 12:
        h += 12
    if ap == "AM" and h == 12:
        h = 0
    return h * 60 + mi


def norm(s):
    return (s or "").strip().upper()


def do_hub(street, city):
    s, c = norm(street), norm(city)
    if "646 MAIN" in s:
        return "646_MAIN_LEW"
    if "618 MAIN" in s:
        return "618_MAIN_LEW"
    if "80 STRAWBERRY" in s:
        return "STRAWBERRY_LEW"
    if "MINOT" in s:
        return "MINOT_AUBURN"
    if "FALCON" in s:
        return "FALCON_LEW"
    if "MANLEY" in s:
        return "MANLEY_AUBURN"
    if "23 CROSS" in s or ("CROSS" in s and "AUBURN" in c):
        return "23_CROSS_AUBURN"
    if "63 BROAD" in s:
        return "63_BROAD_AUBURN"
    if "20 EAST" in s or ("EAST AVE" in s and "LEWISTON" in c):
        return "20_EAST_LEW"
    return f"OTHER|{c[:12]}"


def row_parts(ws, r):
    return [str(ws.cell(r, c).value or "").strip() for c in range(1, 16) if ws.cell(r, c).value]


def parse_groups(ws):
    groups = []
    cur = {"trips": [], "note": ""}
    for r in range(1, (ws.max_row or 0) + 1):
        trip = str(ws.cell(r, 1).value or "").strip()
        meta = str(ws.cell(r, 15).value or "").strip()
        if meta.startswith("__FSGH:"):
            note = trip if not trip.startswith("__FS") else ""
            if cur["trips"]:
                groups.append(cur)
            cur = {"trips": [], "note": note}
            continue
        if not trip or not TRIP.match(trip):
            parts = row_parts(ws, r)
            if parts and not META.match(parts[0]):
                txt = parts[0][:120]
                if "pick" in txt.lower() or "start" in txt.lower() or "together" in txt.lower():
                    cur["note"] = txt
            continue
        pu = parse_time(ws.cell(r, 7).value)
        dof = parse_time(ws.cell(r, 11).value)
        cur["trips"].append(
            {
                "trip": trip,
                "leg": TRIP.match(trip).group(1).upper(),
                "client": norm(ws.cell(r, 3).value),
                "pu_city": norm(ws.cell(r, 5).value),
                "pu": pu,
                "do_hub": do_hub(str(ws.cell(r, 8).value), str(ws.cell(r, 9).value)),
                "do": dof,
            }
        )
    if cur["trips"]:
        groups.append(cur)
    return groups


def main():
    driver_hub = Counter()  # (driver, do_hub) -> groups
    driver_do_time = Counter()  # (driver, do_min rounded) 
    trip_profile_driver = Counter()  # (pu_city, do_hub, do_band) -> driver
    group_do_uniq = Counter()  # groups by dominant DO
    merge_note_groups = 0
    together_notes = 0
    files = 0

    for path in sorted(SCHED.glob("*.xlsx")):
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except Exception:
            continue
        files += 1
        for sn in wb.sheetnames:
            if sn.lower().startswith("template") or sn.lower() in ("reserves", "schedule", "lgtc"):
                continue
            driver = sn.strip()
            groups = parse_groups(wb[sn])
            for g in groups:
                if not g["trips"]:
                    continue
                note = (g.get("note") or "").lower()
                if note:
                    merge_note_groups += 1
                if "togther" in note or "together" in note:
                    together_notes += 1
                dos = [t["do"] for t in g["trips"] if t["do"] and t["do"] > 0]
                hubs = [t["do_hub"] for t in g["trips"] if t["leg"] == "A"]
                if not hubs and g["trips"]:
                    hubs = [g["trips"][0]["do_hub"]]
                dom_hub = Counter(hubs).most_common(1)[0][0] if hubs else "?"
                dom_do = Counter(dos).most_common(1)[0][0] if dos else None
                group_do_uniq[(dom_hub, dom_do)] += 1
                driver_hub[(driver, dom_hub)] += 1
                if dom_do:
                    band = (dom_do // 30) * 30
                    driver_do_time[(driver, band)] += 1
                for t in g["trips"]:
                    if t["leg"] != "A" or not t["pu"] or not dom_do:
                        continue
                    band = (dom_do // 30) * 30
                    trip_profile_driver[(t["pu_city"], dom_hub, band, driver)] += 1
        wb.close()

    # top drivers per (pu_city, hub, do_band)
    profile_best = {}
    by_profile = defaultdict(Counter)
    for key, cnt in trip_profile_driver.items():
        pu, hub, band, driver = key
        by_profile[(pu, hub, band)][driver] += cnt
    for prof, drivers in by_profile.items():
        profile_best[prof] = drivers.most_common(5)

    payload = {
        "files": files,
        "groups_with_notes": merge_note_groups,
        "together_note_groups": together_notes,
        "top_do_waves": [
            {"hub": h, "do_min": d, "groups": c}
            for (h, d), c in group_do_uniq.most_common(25)
        ],
        "driver_hub_leaders": [
            {"driver": d, "hub": h, "groups": c}
            for (d, h), c in driver_hub.most_common(40)
        ],
        "profile_best_drivers_sample": {
            f"{pu}|{hub}|{band}": best
            for (pu, hub, band), best in list(profile_best.items())[:60]
        },
    }
    OUT.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(json.dumps({k: payload[k] for k in ("files", "together_note_groups", "top_do_waves")}, indent=2)[:2000])


if __name__ == "__main__":
    main()
