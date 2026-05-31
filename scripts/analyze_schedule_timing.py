"""Re-read Schedule for 2026 xlsx → person/place early PU + drop timing for Supey BUILD."""
from __future__ import annotations

import json
import re
import statistics
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess
    import sys

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

SCHED_DIR = Path(r"C:\Users\megap\Desktop\Schedule for 2026")
OUT = Path(__file__).resolve().parent / "schedule-timing-analysis.json"

LEG_A = re.compile(r"-A\s*$", re.I)
TIME_RE = re.compile(r"(\d{1,2}):(\d{2})\s*(AM|PM)?", re.I)
LATER_PU = re.compile(r"REQUESTS?\s+LATER", re.I)
CANNOT_DROP = re.compile(r"CANNOT\s+BE\s+DROPPED\s+OFF\s+BEFORE", re.I)


def norm(s: str) -> str:
    return (s or "").strip().upper().replace(".", "").replace(",", "")


def pu_hub(street: str, city: str) -> str:
    s, c = norm(street), norm(city)
    if "MANLEY" in s:
        return "PU_MANLEY"
    if "MINOT" in s and "AUBURN" in c:
        return "PU_MINOT_AUBURN"
    if "FALCON" in s:
        return "PU_FALCON"
    if "646 MAIN" in s:
        return "PU_646_MAIN"
    if "618 MAIN" in s:
        return "PU_618_MAIN"
    if "23 CROSS" in s or ("CROSS ST" in s and "AUBURN" in c):
        return "PU_23_CROSS"
    if "63 BROAD" in s:
        return "PU_63_BROAD"
    if "20 EAST" in s or ("EAST AVE" in s and "LEWISTON" in c):
        return "PU_20_EAST"
    if "10 FALCON" in s:
        return "PU_10_FALCON"
    return ""


def do_hub(street: str, city: str) -> str:
    s, c = norm(street), norm(city)
    if "MANLEY" in s:
        return "MANLEY"
    if "MINOT" in s:
        return "MINOT"
    if "FALCON" in s:
        return "FALCON"
    if "646 MAIN" in s:
        return "646_MAIN"
    if "618 MAIN" in s:
        return "618_MAIN"
    if "23 CROSS" in s or ("CROSS ST" in s and "AUBURN" in c):
        return "23_CROSS"
    if "63 BROAD" in s:
        return "63_BROAD"
    if "20 EAST" in s or ("EAST AVE" in s and "LEWISTON" in c):
        return "20_EAST"
    return "OTHER"


def parse_time(cell) -> int | None:
    if cell is None:
        return None
    if hasattr(cell, "hour"):
        return cell.hour * 60 + cell.minute
    s = str(cell).strip()
    if not s or s.upper() == "NAT":
        return None
    m = TIME_RE.search(s)
    if not m:
        return None
    h, mi = int(m.group(1)), int(m.group(2))
    ap = (m.group(3) or "").upper()
    if ap == "PM" and h != 12:
        h += 12
    if ap == "AM" and h == 12:
        h = 0
    if not ap and h < 12 and "PM" in s.upper():
        h += 12
    return h * 60 + mi


def find_col(header: list[str], *needles: str) -> int | None:
    for i, h in enumerate(header):
        u = (h or "").upper()
        if all(n in u for n in needles):
            return i
    return None


def load_sheet_rows(ws) -> list[dict]:
    """Driver tabs: col0 trip, 2 client, 3-4 PU addr, 6 PU time, 7-8 DO addr, 10 DO time."""
    out = []
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 11:
            continue
        trip = str(row[0] or "").strip()
        if not trip or not LEG_A.search(trip):
            continue
        client = str(row[2] or "").strip().upper()
        if not client:
            continue
        pu_m = parse_time(row[6])
        do_m = parse_time(row[10])
        if pu_m is None or do_m is None or do_m <= pu_m:
            continue
        gap = do_m - pu_m
        note = ""
        for i in range(11, min(len(row), 18)):
            note += " " + str(row[i] or "")
        out.append(
            {
                "client": client,
                "gap": gap,
                "pu_hub": pu_hub(str(row[3] or ""), str(row[4] or "")),
                "do_hub": do_hub(str(row[7] or ""), str(row[8] or "")),
                "later_pu": bool(LATER_PU.search(note)),
                "cannot_drop": bool(CANNOT_DROP.search(note)),
            }
        )
    return out


def p75(vals: list[float]) -> float:
    if not vals:
        return 0.0
    s = sorted(vals)
    i = int(0.75 * (len(s) - 1))
    return float(s[i])


def main() -> None:
    by_client: dict[str, list[int]] = defaultdict(list)
    by_do: dict[str, list[int]] = defaultdict(list)
    by_pu: dict[str, list[int]] = defaultdict(list)
    later_pu: dict[str, int] = defaultdict(int)
    cannot_drop = defaultdict(int)
    files = 0
    trips = 0

    for path in sorted(SCHED_DIR.glob("*.xlsx")):
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception:
            continue
        files += 1
        for name in wb.sheetnames:
            nl = name.lower()
            if nl.startswith("template") or nl == "reserves":
                continue
            ws = wb[name]
            for r in load_sheet_rows(ws):
                trips += 1
                by_client[r["client"]].append(r["gap"])
                if r["do_hub"]:
                    by_do[r["do_hub"]].append(r["gap"])
                if r["pu_hub"]:
                    by_pu[r["pu_hub"]].append(r["gap"])
                if r["later_pu"]:
                    later_pu[r["client"]] += 1
                if r["cannot_drop"]:
                    cannot_drop[r["do_hub"]] += 1
        wb.close()

    client_allow = []
    client_tight = []
    for client, gaps in by_client.items():
        if len(gaps) < 25:
            continue
        p = int(round(p75(gaps)))
        client_allow.append((client, len(gaps), p))
        if p <= 30:
            client_tight.append((client, len(gaps), p))

    client_allow.sort(key=lambda x: (-x[2], -x[1]))
    client_tight.sort(key=lambda x: (x[2], -x[1]))
    no_early = [(c, n) for c, n in sorted(later_pu.items(), key=lambda x: -x[1]) if n >= 3]

    do_hub_p75 = {
        hub: {"n": len(g), "p75": int(round(p75(g)))}
        for hub, g in by_do.items()
        if hub != "OTHER" and len(g) >= 50
    }
    pu_hub_p75 = {
        hub: {"n": len(g), "p75": int(round(p75(g)))}
        for hub, g in by_pu.items()
        if len(g) >= 40
    }

    payload = {
        "files": files,
        "trips": trips,
        "client_early_pu_allowance_p75": client_allow[:50],
        "client_tight_pu_window_p75": client_tight[:30],
        "client_no_early_pu": no_early,
        "do_hub_pu_do_gap_p75": do_hub_p75,
        "pu_hub_pu_do_gap_p75": pu_hub_p75,
        "cannot_drop_notes_by_hub": dict(sorted(cannot_drop.items(), key=lambda x: -x[1])),
    }
    OUT.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"Wrote {OUT} ({files} files, {trips} A-legs)")


if __name__ == "__main__":
    main()
