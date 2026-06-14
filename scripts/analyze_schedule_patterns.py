"""Mine dispatcher patterns from all Schedule for 2026 xlsx files (trips, groups, notes)."""
from __future__ import annotations

import json
import re
import statistics
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess
    import sys

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

SCHED_DIR = Path(r"C:\Users\megap\Desktop\Schedule for 2026")
OUT_JSON = Path(__file__).resolve().parent / "schedule-pattern-analysis.json"
OUT_MD = Path(__file__).resolve().parent / "SCHEDULE_DISPATCHER_PATTERNS.md"

LEG = re.compile(r"-([ABC])\s*$", re.I)
TIME_RE = re.compile(r"(\d{1,2}):(\d{2})\s*(AM|PM)", re.I)
SKIP_SHEETS = {"reserves", "schedule", "lgtc", "template"}

NOTE_PATTERNS = {
    "requests_later_pu": re.compile(r"REQUESTS?\s+LATER", re.I),
    "cannot_drop_before": re.compile(r"CANNOT\s+BE\s+DROPPED\s+OFF\s+BEFORE", re.I),
    "ready_for_pu_at": re.compile(r"READY\s+FOR\s+PU\s+AT", re.I),
    "one_way": re.compile(r"ONE\s+WAY", re.I),
    "door_to_door": re.compile(r"DOOR\s+TO\s+DOOR", re.I),
    "pick_up_at_door": re.compile(r"PICK\s*UP\s+AT\s+DOOR", re.I),
    "cannot_leave_unattended": re.compile(r"CANNOT\s+BE\s+LEFT\s+UNATTENDED|CANT\s+LEAVE\s+UNATTENDED", re.I),
    "needs_assistance": re.compile(r"NEEDS\s+ASSISTANCE", re.I),
    "non_verbal": re.compile(r"NON[- ]VERBAL", re.I),
    "call_first": re.compile(r"CALL\s+(FIRST|BEFORE)", re.I),
    "wheelchair": re.compile(r"WHEEL\s*CHAIR|WC\b", re.I),
    "stretcher": re.compile(r"STRETCHER", re.I),
    "must_be_on_time": re.compile(r"MUST\s+BE\s+ON\s+TIME|ON\s+TIME\s+ONLY", re.I),
    "early_pu_ok": re.compile(r"EARLY\s+(PU|P/U|PIC)", re.I),
    "late_pu_ok": re.compile(r"LATE\s+(PU|P/U|PIC)", re.I),
}


def norm_city(s: str) -> str:
    return (s or "").strip().upper().replace(",", "")


def parse_time(cell) -> int | None:
    if cell is None:
        return None
    if hasattr(cell, "hour") and hasattr(cell, "minute") and not hasattr(cell, "day"):
        return cell.hour * 60 + cell.minute
    if hasattr(cell, "hour") and hasattr(cell, "day"):
        return cell.hour * 60 + cell.minute
    s = str(cell).strip()
    if not s or s.upper() == "NAT":
        return None
    m = TIME_RE.search(s)
    if not m:
        return None
    h, mi = int(m.group(1)), int(m.group(2))
    ap = (m.group(3) or "").upper()
    if ap == "PM" and h != 12:
        h += 12
    if ap == "AM" and h == 12:
        h = 0
    return h * 60 + mi


def is_blank_row(row) -> bool:
    if not row:
        return True
    for c in row[:12]:
        if c is not None and str(c).strip():
            return False
    return True


def row_note(row) -> str:
    parts = []
    for i in range(11, min(len(row), 18)):
        parts.append(str(row[i] or ""))
    return " ".join(parts).strip()


@dataclass
class TripRow:
    trip: str
    leg: str
    client: str
    pu_city: str
    pu_min: int
    do_city: str
    do_min: int
    note: str
    sheet: str
    file: str
    desk_index: int


@dataclass
class GroupBlock:
    trips: list[TripRow] = field(default_factory=list)
    leading_blanks: int = 0
    trailing_note_rows: list[str] = field(default_factory=list)

    @property
    def pu_span(self) -> int | None:
        pus = [t.pu_min for t in self.trips if t.pu_min is not None]
        return max(pus) - min(pus) if len(pus) >= 2 else None

    @property
    def do_span(self) -> int | None:
        dos = [t.do_min for t in self.trips if t.do_min is not None and t.do_min > 0]
        return max(dos) - min(dos) if len(dos) >= 2 else None

    @property
    def pu_cities(self) -> set[str]:
        return {t.pu_city for t in self.trips if t.pu_city}

    @property
    def do_cities(self) -> set[str]:
        return {t.do_city for t in self.trips if t.do_city}


def parse_trip_row(row, sheet: str, file: str, desk_index: int) -> TripRow | None:
    if not row or len(row) < 11:
        return None
    trip = str(row[0] or "").strip()
    m = LEG.search(trip)
    if not m:
        return None
    pu = parse_time(row[6])
    dof = parse_time(row[10])
    client = str(row[2] or "").strip().upper()
    if not client:
        return None
    return TripRow(
        trip=trip,
        leg=m.group(1).upper(),
        client=client,
        pu_city=norm_city(str(row[4] or "")),
        pu_min=pu,
        do_city=norm_city(str(row[8] or "")),
        do_min=dof,
        note=row_note(row),
        sheet=sheet,
        file=file,
        desk_index=desk_index,
    )


def split_groups(rows: list) -> list[GroupBlock]:
    groups: list[GroupBlock] = []
    current = GroupBlock()
    blank_run = 0
    desk = 0
    for row in rows:
        if is_blank_row(row):
            blank_run += 1
            if current.trips:
                current.trailing_note_rows = []
                groups.append(current)
                current = GroupBlock(leading_blanks=blank_run)
            continue
        t = parse_trip_row(row, "", "", desk)
        if not t:
            continue
        if blank_run and not current.trips:
            current.leading_blanks = blank_run
        blank_run = 0
        current.trips.append(t)
        desk += 1
    if current.trips:
        groups.append(current)
    return groups


def pct(n: int, d: int) -> float:
    return round(100.0 * n / d, 1) if d else 0.0


def p75(vals: list[float | int]) -> float:
    if not vals:
        return 0.0
    s = sorted(vals)
    i = int(0.75 * (len(s) - 1))
    return float(s[i])


def main() -> None:
    stats = {
        "files": 0,
        "driver_sheets": 0,
        "trips": 0,
        "groups": 0,
        "blank_row_group_breaks": 0,
        "inferred_90min_would_split": 0,
    }

    group_sizes: Counter[int] = Counter()
    pu_spans: list[int] = []
    do_spans: list[int] = []
    pu_cities_per_group: list[int] = []
    do_cities_per_group: list[int] = []
    desk_vs_pu_reorders: Counter[str] = Counter()
    inter_group_gaps: list[int] = []  # prior last event to next first PU
    same_do_time_batches: Counter[int] = Counter()
    a_leg_800_do_waves: Counter[int] = Counter()
    note_hits: Counter[str] = Counter()
    note_examples: dict[str, list[str]] = defaultdict(list)
    leg_mix_in_group: Counter[str] = Counter()
    concurrent_pu_14: Counter[int] = Counter()
    concurrent_pu_29: Counter[int] = Counter()
    concurrent_pu_120: Counter[int] = Counter()
    group_first_pu_vs_min_pu: list[int] = []  # desk first trip PU minus min PU in group
    split_client_ab_same_day: Counter[str] = Counter()
    examples_multi_city = []
    examples_wide_pu_span = []
    examples_inter_group_tight = []
    examples_desk_reorder = []
    driver_group_counts: Counter[str] = Counter()

    for path in sorted(SCHED_DIR.glob("*.xlsx")):
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception:
            continue
        stats["files"] += 1
        fname = path.name

        for sheet_name in wb.sheetnames:
            if sheet_name.strip().lower() in SKIP_SHEETS or sheet_name.lower().startswith("template"):
                continue
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            stats["driver_sheets"] += 1
            driver_group_counts[sheet_name.strip()] += 1

            groups = split_groups(rows)
            stats["groups"] += len(groups)
            stats["blank_row_group_breaks"] += sum(1 for g in groups if g.leading_blanks > 0)

            prev_last_do = None
            prev_last_pu = None
            for gi, g in enumerate(groups):
                n = len(g.trips)
                group_sizes[n] += 1
                stats["trips"] += n
                driver_group_counts[sheet_name.strip() + ":groups"] += 1

                if g.pu_span is not None:
                    pu_spans.append(g.pu_span)
                if g.do_span is not None:
                    do_spans.append(g.do_span)
                pu_cities_per_group.append(len(g.pu_cities))
                do_cities_per_group.append(len(g.do_cities))

                if len(g.pu_cities) >= 3 and len(examples_multi_city) < 8:
                    examples_multi_city.append(
                        {
                            "file": fname,
                            "driver": sheet_name,
                            "group": gi + 1,
                            "pu_cities": sorted(g.pu_cities),
                            "pu_span_min": g.pu_span,
                            "trips": [(t.trip, t.pu_min, t.pu_city) for t in g.trips[:6]],
                        }
                    )

                if g.pu_span and g.pu_span >= 55 and len(examples_wide_pu_span) < 8:
                    examples_wide_pu_span.append(
                        {
                            "file": fname,
                            "driver": sheet_name,
                            "pu_span": g.pu_span,
                            "trips": [(t.trip, t.pu_min, t.pu_city) for t in g.trips],
                        }
                    )

                pus = sorted([t.pu_min for t in g.trips if t.pu_min is not None])
                if len(pus) >= 2:
                    concurrent_pu_14[sum(1 for i, p in enumerate(pus) if i == 0 or p - pus[i - 1] <= 14)] += 1
                    concurrent_pu_29[sum(1 for i, p in enumerate(pus) if i == 0 or p - pus[i - 1] <= 29)] += 1
                    concurrent_pu_120[sum(1 for i, p in enumerate(pus) if i == 0 or p - pus[i - 1] <= 120)] += 1

                # desk order vs PU order
                with_pu = [t for t in g.trips if t.pu_min is not None]
                if len(with_pu) >= 2:
                    desk_order = [t.pu_min for t in with_pu]
                    pu_sorted = sorted(desk_order)
                    if desk_order != pu_sorted:
                        desk_vs_pu_reorders["reordered"] += 1
                        if len(examples_desk_reorder) < 6:
                            examples_desk_reorder.append(
                                {
                                    "file": fname,
                                    "driver": sheet_name,
                                    "desk": [(t.trip, t.pu_min, t.pu_city) for t in with_pu],
                                    "pu_sorted": pu_sorted,
                                }
                            )
                    else:
                        desk_vs_pu_reorders["pu_chronological"] += 1

                # DO batching (same DO minute)
                dos = [t.do_min for t in g.trips if t.do_min and t.do_min > 0]
                if dos:
                    top_do = Counter(dos).most_common(1)[0]
                    if top_do[1] >= 2:
                        same_do_time_batches[top_do[1]] += 1
                    if all(t.leg == "A" for t in g.trips):
                        eight_am = sum(1 for d in dos if d == 8 * 60)
                        if eight_am >= 2:
                            a_leg_800_do_waves[eight_am] += 1

                legs = Counter(t.leg for t in g.trips)
                leg_mix_in_group[",".join(sorted(legs.keys()))] += 1

                min_pu = min((t.pu_min for t in g.trips if t.pu_min is not None), default=None)
                first_desk_pu = next((t.pu_min for t in g.trips if t.pu_min is not None), None)
                if min_pu is not None and first_desk_pu is not None:
                    group_first_pu_vs_min_pu.append(first_desk_pu - min_pu)

                for t in g.trips:
                    for key, rx in NOTE_PATTERNS.items():
                        if rx.search(t.note):
                            note_hits[key] += 1
                            if len(note_examples[key]) < 3:
                                note_examples[key].append(t.note[:120])

                # Inter-group: use last DO (or PU) of prior group to first PU of this group
                first_pu = next((t.pu_min for t in g.trips if t.pu_min is not None), None)
                if prev_last_do is not None and first_pu is not None and prev_last_do > 0:
                    inter_group_gaps.append(first_pu - prev_last_do)
                    if 0 <= first_pu - prev_last_do <= 20 and len(examples_inter_group_tight) < 8:
                        examples_inter_group_tight.append(
                            {
                                "file": fname,
                                "driver": sheet_name,
                                "gap_min": first_pu - prev_last_do,
                                "prev_do": prev_last_do,
                                "next_pu": first_pu,
                            }
                        )

                last_do = max((t.do_min for t in g.trips if t.do_min and t.do_min > 0), default=None)
                last_pu = max((t.pu_min for t in g.trips if t.pu_min is not None), default=None)
                prev_last_do = last_do if last_do else last_pu
                prev_last_pu = last_pu

                # Would 90-min rule split this group?
                prev = None
                for t in g.trips:
                    if t.pu_min is None:
                        continue
                    if prev is not None and t.pu_min - prev > 90:
                        stats["inferred_90min_would_split"] += 1
                        break
                    prev = t.pu_min

            # A+B same client same driver
            clients_a = {t.client for t in sum([g.trips for g in groups], []) if t.leg == "A"}
            clients_b = {t.client for t in sum([g.trips for g in groups], []) if t.leg == "B"}
            for c in clients_a & clients_b:
                split_client_ab_same_day[c] += 1

        wb.close()

    # Summarize
    total_groups = stats["groups"]
    multi_city_3plus = sum(1 for x in pu_cities_per_group if x >= 3)
    multi_city_2plus = sum(1 for x in pu_cities_per_group if x >= 2)

    summary = {
        "stats": stats,
        "group_size_distribution": dict(group_sizes.most_common(12)),
        "pu_span_minutes_p75": int(round(p75(pu_spans))),
        "pu_span_minutes_max": max(pu_spans) if pu_spans else 0,
        "do_span_minutes_p75": int(round(p75(do_spans))),
        "groups_with_2plus_pu_cities_pct": pct(multi_city_2plus, total_groups),
        "groups_with_3plus_pu_cities_pct": pct(multi_city_3plus, total_groups),
        "pu_cities_per_group_p75": p75(pu_cities_per_group),
        "desk_order_vs_pu_chronological": dict(desk_vs_pu_reorders),
        "inter_group_gap_minutes_p75": int(round(p75(inter_group_gaps))),
        "inter_group_gap_negative_count": sum(1 for g in inter_group_gaps if g < 0),
        "inter_group_gap_under_30_count": sum(1 for g in inter_group_gaps if 0 <= g < 30),
        "same_do_time_batch_sizes": dict(same_do_time_batches.most_common(8)),
        "a_leg_8am_do_wave_sizes": dict(a_leg_800_do_waves.most_common(8)),
        "note_pattern_counts": dict(note_hits.most_common()),
        "note_examples": dict(note_examples),
        "leg_mix_in_group": dict(leg_mix_in_group.most_common(10)),
        "concurrent_pu_within_14_distribution": dict(concurrent_pu_14.most_common()),
        "concurrent_pu_within_29_distribution": dict(concurrent_pu_29.most_common()),
        "concurrent_pu_within_120_distribution": dict(concurrent_pu_120.most_common()),
        "desk_first_trip_pu_minus_group_min_pu": {
            "p75": int(round(p75(group_first_pu_vs_min_pu))),
            "nonzero_pct": pct(sum(1 for x in group_first_pu_vs_min_pu if x > 0), len(group_first_pu_vs_min_pu)),
        },
        "clients_with_a_and_b_same_driver_top20": split_client_ab_same_day.most_common(20),
        "examples_multi_city_groups": examples_multi_city,
        "examples_wide_pu_span": examples_wide_pu_span,
        "examples_inter_group_tight": examples_inter_group_tight,
        "examples_desk_reorder": examples_desk_reorder,
    }

    OUT_JSON.write_text(json.dumps(summary, indent=2, ensure_ascii=False),encoding="utf-8")

    md = render_md(summary)
    OUT_MD.write_text(md,encoding="utf-8")
    print(f"Analyzed {stats['files']} files, {stats['trips']} trips, {stats['groups']} groups")
    print(f"Wrote {OUT_JSON.name} and {OUT_MD.name}")


def render_md(s: dict) -> str:
    st = s["stats"]
    lines = [
        "# Dispatcher patterns from real schedules",
        "",
        f"Source: `{SCHED_DIR}` — **{st['files']}** workbook days, **{st['trips']:,}** trips, **{st['groups']:,}** route groups.",
        "",
        "## How groups are drawn on the sheet",
        "",
        f"- **Blank row = new group** in **{pct(st['blank_row_group_breaks'], st['groups']):.0f}%** of groups (explicit route break).",
        f"- Groups with **90+ min pickup gaps** inside one block (would split if inferred from time only): **{st['inferred_90min_would_split']}** cases — dispatchers often **keep one group** anyway.",
        f"- Typical group size: {', '.join(f'{k} trips×{v}' for k,v in list(s['group_size_distribution'].items())[:6])}.",
        "",
        "## Geography — dispatchers routinely batch distant towns",
        "",
        f"- **{s['groups_with_2plus_pu_cities_pct']}%** of groups pick up in **2+ cities**.",
        f"- **{s['groups_with_3plus_pu_cities_pct']}%** pick up in **3+ cities** (Poland + Mechanic Falls + Lewiston in one morning batch is normal).",
        f"- PU span inside a group: **p75 {s['pu_span_minutes_p75']} min**, max **{s['pu_span_minutes_max']} min**.",
        f"- Desk order ≠ PU-time order in **{s['desk_order_vs_pu_chronological'].get('reordered', 0)}** groups vs chronological **{s['desk_order_vs_pu_chronological'].get('pu_chronological', 0)}** — **route order on sheet is intentional**, not sort-by-PU.",
        "",
        "### Example multi-city groups",
        "",
    ]
    for ex in s.get("examples_multi_city_groups", [])[:5]:
        lines.append(f"- **{ex['driver']}** ({ex['file']}): PU cities {ex['pu_cities']}, span {ex['pu_span_min']} min")

    lines += [
        "",
        "## Timing tricks",
        "",
        f"- **Shared drop time batches**: {s.get('same_do_time_batch_sizes', {})} — many A-legs share **8:00 AM DO** in one group.",
        f"- **8:00 AM DO waves** (A-leg only): {s.get('a_leg_8am_do_wave_sizes', {})}.",
        f"- First trip on sheet vs earliest PU in group: **{s['desk_first_trip_pu_minus_group_min_pu']['nonzero_pct']}%** of groups start desk list **before** the earliest scheduled PU (early pickup / different first stop).",
        f"- Inter-group: prior DO → next PU p75 **{s['inter_group_gap_minutes_p75']} min**; **{s['inter_group_gap_under_30_count']}** transitions under 30 min; **{s['inter_group_gap_negative_count']}** where next PU is *before* prior DO (overlap batches / separate groups on same driver).",
        "",
        "## Notes in trip comments (col 13+)",
        "",
    ]
    for key, count in s.get("note_pattern_counts", {}).items():
        ex = (s.get("note_examples", {}).get(key) or [""])[0]
        lines.append(f"- **{key.replace('_', ' ')}**: {count}× — e.g. `{ex[:90]}…`" if len(ex) > 90 else f"- **{key.replace('_', ' ')}**: {count}× — e.g. `{ex}`")

    lines += [
        "",
        "## A + B same driver same day",
        "",
        "Clients with both A-leg (morning) and B-leg (return) on one driver sheet:",
        "",
    ]
    for client, n in s.get("clients_with_a_and_b_same_driver_top20", [])[:10]:
        lines.append(f"- {client}: {n} schedule-days")

    lines += [
        "",
        "## Implications for suggest-driver",
        "",
        "1. **Do not require PU-time sort** — honor desk/route order; BUILD uses OSRM tour order.",
        "2. **Multi-city morning batches are valid** — reject merge only when OSRM tour fails, not when cities look far apart on a map.",
        "3. **Blank row groups** — prefer `NewGroupAfterGroup` with explicit gaps; allow overlap transitions when inter-group timing shows next PU before prior DO (separate batches).",
        "4. **Comment-driven rules** — `REQUESTS LATER PU` = no early PU; `CANNOT BE DROPPED OFF BEFORE` = program DO floor; `READY FOR PU AT` = hard PU floor.",
        "5. **8:00 DO clusters** — scoring should prefer merging into groups that share DO anchor times.",
        "",
        "Regenerate: `python scripts/analyze_schedule_patterns.py`",
    ]
    return "\n".join(lines) + "\n"


if __name__ == "__main__":
    main()
