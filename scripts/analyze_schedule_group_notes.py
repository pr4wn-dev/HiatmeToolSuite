"""Extract dispatcher *schedule notes* (gap/header rows), not trip comment fields."""
from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess
    import sys

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

SCHED_DIR = Path(r"C:\Users\megap\Desktop\Schedule for 2026")
OUT_JSON = Path(__file__).resolve().parent / "schedule-group-notes-analysis.json"
OUT_MD = Path(__file__).resolve().parent / "SCHEDULE_GROUP_NOTES.md"

TRIP = re.compile(r"^1-\d+-[ABC]$", re.I)
META = re.compile(r"^__FS", re.I)
NOISE = re.compile(
    r"^(x+|z+|w+|r+|a+|\*+|\.+|xxxx|xxx|zzzz|placeholder)$",
    re.I,
)

CATEGORIES = [
    ("pick_up_together", re.compile(r"pick\s*up\s*to(g|ge)ther|togther|together", re.I)),
    ("pick_up_in_order", re.compile(r"pick\s*up.*in\s*order|in\s*order", re.I)),
    ("round_trip", re.compile(r"round\s*trip", re.I)),
    ("start_time", re.compile(r"\d{1,2}[:.]?\d{0,2}\s*(am|pm)\s*start|start\s*at|pick\s*up\s*at|pickup\s*at|\d{3,4}\s*am\s*start", re.I)),
    ("early_pickup_time", re.compile(r"6:50|6:45|645|650|7:00|7am|6:40", re.I)),
    ("come_to_office", re.compile(r"come\s*to\s*office|office\s*first", re.I)),
    ("get_client", re.compile(r"get\s+\w+|pick\s*up\s+\w+.*first", re.I)),
    ("will_call", re.compile(r"will\s*call", re.I)),
    ("usual_set", re.compile(r"usual\s*set", re.I)),
    ("dont_be_late", re.compile(r"don.?t\s*be\s*late|be\s*on\s*time", re.I)),
    ("one_way", re.compile(r"one\s*way", re.I)),
    ("wait", re.compile(r"\bwait\b|hold", re.I)),
    ("drop_order", re.compile(r"drop.*first|drop.*order|drop\s*off\s*order", re.I)),
]


def row_values(ws, row_idx: int) -> list[str]:
    out = []
    for c in range(1, 16):
        v = ws.cell(row_idx, c).value
        if v is not None and str(v).strip():
            out.append(str(v).strip())
    return out


def normalize_note(text: str) -> str:
    return " ".join(text.split()).strip()


def categorize(note: str) -> list[str]:
    hits = []
    for name, rx in CATEGORIES:
        if rx.search(note):
            hits.append(name)
    return hits or ["other"]


def main() -> None:
    raw_notes: Counter[str] = Counter()
    by_category: Counter[str] = Counter()
    examples: dict[str, list] = defaultdict(list)
    by_driver: Counter[str] = Counter()
    files = 0
    note_rows = 0

    for path in sorted(SCHED_DIR.glob("*.xlsx")):
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except Exception:
            continue
        files += 1

        for sheet in wb.sheetnames:
            sl = sheet.lower()
            if sl.startswith("template") or sl in ("reserves", "schedule", "lgtc"):
                continue

            ws = wb[sheet]
            for r in range(1, (ws.max_row or 0) + 1):
                parts = row_values(ws, r)
                if not parts:
                    continue

                # Group header metadata only — skip unless col A has user text
                if any(META.match(p) for p in parts):
                    user = [p for p in parts if not META.match(p)]
                    if not user:
                        continue
                    note = normalize_note(user[0])
                else:
                    if TRIP.match(parts[0]):
                        continue
                    note = normalize_note(parts[0])

                if len(note) < 3 or NOISE.match(note.replace(" ", "")):
                    continue

                note_rows += 1
                raw_notes[note.lower()] += 1
                by_driver[sheet.strip()] += 1

                cats = categorize(note)
                for cat in cats:
                    by_category[cat] += 1
                    if len(examples[cat]) < 5:
                        examples[cat].append(
                            {
                                "note": note,
                                "file": path.name,
                                "driver": sheet.strip(),
                                "row": r,
                            }
                        )

        wb.close()

    # Merge similar raw notes for display
    top_raw = raw_notes.most_common(60)

    payload = {
        "files_scanned": files,
        "schedule_note_rows": note_rows,
        "unique_notes": len(raw_notes),
        "by_category": dict(by_category.most_common()),
        "top_raw_notes": [{"note": n, "count": c} for n, c in top_raw],
        "examples_by_category": dict(examples),
        "drivers_with_most_notes": by_driver.most_common(15),
    }

    OUT_JSON.write_text(json.dumps(payload, indent=2, ensure_ascii=False),encoding="utf-8")
    OUT_MD.write_text(render_md(payload),encoding="utf-8")
    print(f"{files} files, {note_rows} schedule note rows, {len(raw_notes)} unique")
    print(f"Wrote {OUT_MD.name}")


def render_md(p: dict) -> str:
    lines = [
        "# Schedule group / route notes (dispatcher-written)",
        "",
        f"Scanned **{p['files_scanned']}** xlsx days — **{p['schedule_note_rows']}** note rows on driver sheets",
        f"(blank/gap/header rows — **not** Modivcare trip comment columns).",
        "",
        "## What these notes are",
        "",
        "Dispatchers type short instructions on the **colored group bar row** or **gap row** between batches:",
        "start times, pick-up together, round trip, get a specific client first, etc.",
        "",
        "## Categories (from note text)",
        "",
    ]
    for cat, count in sorted(p["by_category"].items(), key=lambda x: -x[1]):
        lines.append(f"### {cat.replace('_', ' ')} ({count} hits)")
        for ex in p.get("examples_by_category", {}).get(cat, [])[:4]:
            lines.append(
                f"- \"{ex['note']}\" — {ex['driver']}, {ex['file']}, row {ex['row']}"
            )
        lines.append("")

    lines += [
        "## Most common exact note text",
        "",
    ]
    for item in p.get("top_raw_notes", [])[:35]:
        lines.append(f"- **{item['count']}×** \"{item['note']}\"")

    lines += [
        "",
        "## What suggest-driver / BUILD should do with these",
        "",
        "- **pick up together / in order** — trips in the following group share one OSRM tour; desk order matters.",
        "- **650 am start / 6:50 pick up** — driver clock anchor for the group, not first trip's Modivcare PU time.",
        "- **GET [client]** — first stop in tour may differ from earliest scheduled PU.",
        "- **ROUND TRIP** — A + B same driver; keep returns paired.",
        "- **COME TO OFFICE** — deadhead starts at office, not home.",
        "- **WILL CALL** — flex PU; don't treat scheduled PU as hard floor.",
        "- **usual set** — match weekday template grouping for that driver.",
        "",
        "Regenerate: `python scripts/analyze_schedule_group_notes.py`",
    ]
    return "\n".join(lines) + "\n"


if __name__ == "__main__":
    main()
